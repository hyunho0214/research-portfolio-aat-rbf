"""Panama electricity demand loading helpers."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


def find_longest_demand_sheet(path: Path) -> str:
    """Select the longest Excel sheet that contains a DEMAND column."""

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        candidates: list[tuple[int, str]] = []
        for ws in wb.worksheets:
            header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            if "datetime" in header and "DEMAND" in header:
                candidates.append((ws.max_row, ws.title))
        if not candidates:
            raise ValueError(f"No sheet with datetime and DEMAND columns found in {path}.")
        return max(candidates)[1]
    finally:
        wb.close()


def load_excel_demand_series(
    path: Path,
    *,
    sheet_name: str | None = None,
    start_date: str = "2016-01-01",
) -> pd.DataFrame:
    """Load hourly actual demand from the selected Excel sheet."""

    selected_sheet = sheet_name or find_longest_demand_sheet(path)
    df = pd.read_excel(path, sheet_name=selected_sheet, usecols=["datetime", "DEMAND"])
    df["datetime"] = pd.to_datetime(df["datetime"])
    df = df.sort_values("datetime").dropna(subset=["DEMAND"])
    if start_date:
        df = df[df["datetime"] >= pd.Timestamp(start_date)]
    df = df.rename(columns={"DEMAND": "demand"})
    df.attrs["sheet_name"] = selected_sheet
    return df.reset_index(drop=True)


def load_csv_forecast_series(path: Path, *, start_date: str = "2016-01-01") -> pd.DataFrame:
    """Load hourly pre-dispatch forecast data for optional comparison runs."""

    df = pd.read_csv(path)
    df["datetime"] = pd.to_datetime(df["datetime"])
    df = df.sort_values("datetime").dropna(subset=["load_forecast"])
    if start_date:
        df = df[df["datetime"] >= pd.Timestamp(start_date)]
    df = df.rename(columns={"load_forecast": "demand"})
    df.attrs["sheet_name"] = "weekly pre-dispatch forecast.csv"
    return df[["datetime", "demand"]].reset_index(drop=True)
